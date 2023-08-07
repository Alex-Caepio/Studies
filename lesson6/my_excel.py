import openpyxl

excel_file_path = 'excel.xlsx'

workbook = openpyxl.load_workbook(excel_file_path)

sheet = workbook.active

for row in sheet.iter_rows(values_only=True):
    print(row)

# write to excel

excel_file_path = 'excel.xlsx'

workbook = openpyxl.Workbook()

sheet = workbook.active

sheet['A3'] = 'Hello'

sheet.append([1, 2, 3])

print(sheet['A3'].value)

workbook.save(excel_file_path)

excel_file_path = 'excel.xlsx'

workbook = openpyxl.Workbook()

sheet = workbook.active

data = [
    ['Name', 'Age', 'Job'],
    ['Bob', 25, 'Dev'],
    ['Alice', 28, 'Dev'],
    ['Charlie', 30, 'Manager']
]

for row in data:
    sheet.append(row)

workbook.save(excel_file_path)
